# coding=utf-8
import json,requests
import openpyxl
import sys


def login(username,password):
    url_login="https://api.zoomeye.org/user/login"
    data={
        "username": "",
        "password": ""
    }
    data["username"]=username
    data["password"]=password
    data=json.dumps(data)
    r=requests.post(url=url_login,data=data)
    return json.loads(r.content)['access_token']

def GetResidual(token):
    url="https://api.zoomeye.org/resources-info"
    headers={'Authorization':'JWT ' + token}
    r=requests.get(url=url,headers=headers)
    datas=json.loads(r.content)
    print("剩余搜索次数: {}".format(datas['resources']['search']))

def Search(token,search,page):
    headers = {'Authorization': 'JWT ' + token}
    # 创建一个工作簿
    workbook_new = openpyxl.Workbook()
    # 创建一个表单
    sheet_new = workbook_new.create_sheet('info')
    id=0
    try:
        for i in range(0,int(page)):
            a=str(i)
            url = "https://api.zoomeye.org/web/search?query={}&page={}".format(search, a)
            r=requests.get(url=url,headers=headers)
            data = json.loads(r.content)['matches']
            json_len = len(data)

            for item in range(0, json_len):
                try:
                    # 表格仅获取ip和site信息，根据需要修改代码添加存储字段
                    #从json取IP
                    if(data[item]['ip'][0]!="*"):
                        sheet_new.cell(row=id, column=1, value=data[item]['ip'][0])
                    #从json取site
                    sheet_new.cell(row=id, column=2, value=data[item]['site'])
                except Exception:
                    print("输出错误+1")
                    pass
                id=id+1
    except Exception:
        pass
    workbook_new.save('test.xlsx')

if __name__== "__main__":
    arch=sys.argv[1]
    if(arch=="-q"):
        # 进行查询
        username = sys.argv[2]
        password = sys.argv[3]
        # 登录获取cookie
        token = login(username, password)
        page = sys.argv[4]
        search = sys.argv[5]
        Search(token,search,page)
    elif(arch=="-c"):
        # 获取剩余查询次数
        username = sys.argv[2]
        password = sys.argv[3]
        # 登录获取cookie
        token = login(username, password)
        GetResidual(token)
    else:
        print("查看帮助：python zoomeyes.py -h")
        print("查询剩余次数：python zoomeyes.py -c 账号 密码")
        print("查询：python zoomeyes.py -q 账号 密码 查询页数 查询语句")
        print("eg:python zoomeyes.py -q 123456@qq.com 123456 2 \"app:Microsoft Exchange imapd\"")
        print("***查询结果输出本地test.xlsx文件info表***")
        print("表格仅获取ip和site信息，根据需要修改代码添加存储字段")